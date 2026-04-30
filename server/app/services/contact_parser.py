"""Stream-parse contact datasheets from CSV, XLSX, or JSON payloads.

Yields normalized dicts. Never loads the entire file into memory for CSV/JSON;
XLSX is read via openpyxl's iterator mode.
"""
from __future__ import annotations

import csv
import io
import json
import re
from typing import Iterable, Iterator

from openpyxl import load_workbook

_EMAIL_RE = re.compile(r"^[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}$", re.IGNORECASE)
_PHONE_RE = re.compile(r"[^\d+]")

_FIELD_ALIASES = {
    "email": {"email", "email_address", "e-mail", "mail"},
    "first_name": {"first_name", "firstname", "given_name", "first", "fname"},
    "last_name": {"last_name", "lastname", "surname", "family_name", "last", "lname"},
    "phone": {"phone", "phone_number", "mobile", "telephone", "tel"},
    "company": {"company", "organization", "org", "employer"},
    "company_url": {"company_url", "website", "url", "domain"},
    "job_title": {"job_title", "title", "role", "position"},
    "industry": {"industry", "sector", "vertical"},
    "tz": {"tz", "timezone", "time_zone"},
    "source_tag": {"source_tag", "source", "list_name", "origin"},
    "confidence_score": {"confidence_score", "confidence", "score", "quality"},
}

# Columns present in the downloadable sample template (and the canonical
# order they ship in). Drives both the CSV template generator and the
# `validate_headers()` completeness report.
TEMPLATE_COLUMNS = [
    "email",
    "first_name",
    "last_name",
    "company",
    "job_title",
    "phone",
    "company_url",
    "industry",
    "tz",
    "source_tag",
]

# Only email is strictly required. Everything else is optional and merged
# into refined_contacts via precedence rules.
REQUIRED_COLUMNS = {"email"}


def validate_headers(headers: list[str]) -> dict[str, object]:
    """Classify header row against the canonical template.

    Returns:
        {
          "matched": {canonical: original_header, ...},
          "missing_required": [canonical, ...],
          "unknown": [original_header, ...]  # kept as custom_fields at ingest
        }
    """
    matched: dict[str, str] = {}
    unknown: list[str] = []
    for h in headers:
        if not h:
            continue
        canonical = _canonical_field(h)
        if canonical is None:
            unknown.append(h.strip())
        else:
            matched.setdefault(canonical, h.strip())
    missing_required = sorted(REQUIRED_COLUMNS - matched.keys())
    return {
        "matched": matched,
        "missing_required": missing_required,
        "unknown": unknown,
    }


def _canonical_field(header: str) -> str | None:
    key = header.strip().lower().replace(" ", "_")
    for canonical, aliases in _FIELD_ALIASES.items():
        if key in aliases:
            return canonical
    return None


def _normalize_phone(raw: str) -> str:
    if not raw:
        return ""
    cleaned = _PHONE_RE.sub("", raw)
    if cleaned and not cleaned.startswith("+"):
        cleaned = "+" + cleaned
    return cleaned


def _normalize_row(raw_row: dict[str, str]) -> dict[str, object] | None:
    """Back-compat wrapper that discards the drop reason."""
    normalized, _reason = _normalize_row_with_reason(raw_row)
    return normalized


def _normalize_row_with_reason(
    raw_row: dict[str, str],
) -> tuple[dict[str, object] | None, str | None]:
    """Normalize a raw header->value row.

    Returns ``(row, None)`` on success or ``(None, reason)`` when the row
    has to be dropped. Reason is one of:
        - 'no_email'        : no recognized email column OR blank value
        - 'invalid_email'   : value present but fails regex
        - 'custom_only'     : every header was unknown (no canonical match)
    """
    out: dict[str, object] = {"custom_fields": {}}
    any_canonical = False
    for header, value in raw_row.items():
        if header is None:
            continue
        canonical = _canonical_field(header)
        str_value = "" if value is None else str(value).strip()
        if canonical is None:
            if str_value:
                out["custom_fields"][header.strip()] = str_value
            continue
        any_canonical = True
        if canonical == "email":
            out["email"] = str_value.lower()
        elif canonical == "phone":
            out["phone"] = _normalize_phone(str_value)
        elif canonical == "confidence_score":
            # Numeric cast — downstream upsert tests isinstance(int, float).
            # Without this, string values silently fail the type check and
            # never make it into refined_contacts.confidence_score.
            if str_value:
                try:
                    out["confidence_score"] = float(str_value)
                except ValueError:
                    pass  # malformed numeric — drop field but keep row
        else:
            out[canonical] = str_value

    email = out.get("email")
    if not email:
        return None, "no_email" if any_canonical else "custom_only"
    if not isinstance(email, str) or not _EMAIL_RE.match(email):
        return None, "invalid_email"
    # Flag unnamed rows so the ingest counter can surface them. Row is
    # still accepted — a lead is better than nothing — but the template
    # renderer will substitute a greeting fallback and the UI warns if
    # the unnamed share is large.
    if not out.get("first_name") and not out.get("last_name"):
        out["_missing_name"] = True
    return out, None


def parse_csv(stream: io.IOBase) -> Iterator[dict[str, object]]:
    reader = csv.DictReader(io.TextIOWrapper(stream, encoding="utf-8", newline=""))
    for raw in reader:
        normalized = _normalize_row(raw)
        if normalized is not None:
            yield normalized


def parse_json(stream: io.IOBase) -> Iterator[dict[str, object]]:
    """Stream-parse a JSON contact file item-by-item using ijson.

    Supports two shapes:
      - top-level array: [ {...}, {...}, ... ]          -> items at "item"
      - envelope object: { "contacts": [ {...}, ... ] } -> items at "contacts.item"

    A 500 MB file never materializes in memory — only one row at a time.
    Phase 1 change #6 in docs/3phase_implementation.md.
    """
    import ijson  # lazy import: only this parser path needs it

    # Detect shape from the first non-whitespace byte without consuming
    # more than one char — `ijson` handles everything after this probe.
    head = stream.read(1)
    # Put it back at the front for ijson to re-read.
    remainder = stream.read()
    combined = io.BytesIO(head + remainder)

    if head == b"[":
        prefix = "item"
    elif head == b"{":
        prefix = "contacts.item"
    else:
        # Unknown shape; stay safe and yield nothing rather than OOM on a
        # full json.loads() of an arbitrary blob.
        return

    for raw in ijson.items(combined, prefix):
        if not isinstance(raw, dict):
            continue
        normalized = _normalize_row(
            {k: ("" if v is None else str(v)) for k, v in raw.items()}
        )
        if normalized is not None:
            yield normalized


def parse_xlsx(raw_bytes: bytes) -> Iterator[dict[str, object]]:
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return
    headers = [str(h) if h is not None else "" for h in header_row]
    for row in rows_iter:
        raw = {headers[i]: (row[i] if i < len(row) else "") for i in range(len(headers))}
        normalized = _normalize_row({k: ("" if v is None else str(v)) for k, v in raw.items()})
        if normalized is not None:
            yield normalized


def parse_stream(filename: str, content: bytes) -> Iterator[dict[str, object]]:
    name = filename.lower()
    if name.endswith(".csv"):
        yield from parse_csv(io.BytesIO(content))
    elif name.endswith(".json"):
        yield from parse_json(io.BytesIO(content))
    elif name.endswith(".xlsx"):
        yield from parse_xlsx(content)
    else:
        raise ValueError(f"Unsupported file type: {filename}")


def peek_headers(filename: str, content: bytes) -> list[str]:
    """Return the header row without materializing the whole file.

    Used by the extraction agent to run `validate_headers()` before
    streaming rows, so we can fail fast with a structured error when
    the required 'email' column is missing instead of silently dropping
    every row.
    """
    name = filename.lower()
    if name.endswith(".csv"):
        # Read the first logical CSV line — respects quoted fields.
        stream = io.TextIOWrapper(io.BytesIO(content), encoding="utf-8", newline="")
        reader = csv.reader(stream)
        try:
            return [str(h).strip() for h in next(reader)]
        except StopIteration:
            return []
    if name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []
        try:
            header_row = next(ws.iter_rows(values_only=True))
        except StopIteration:
            return []
        return [str(h).strip() if h is not None else "" for h in header_row]
    if name.endswith(".json"):
        # JSON is self-describing per-object. Peek at the first object's keys.
        import ijson

        bio = io.BytesIO(content)
        head = bio.read(1)
        bio.seek(0)
        if head == b"[":
            prefix = "item"
        elif head == b"{":
            prefix = "contacts.item"
        else:
            return []
        for raw in ijson.items(bio, prefix):
            if isinstance(raw, dict):
                return [str(k).strip() for k in raw.keys()]
            break
        return []
    raise ValueError(f"Unsupported file type: {filename}")


def parse_stream_with_stats(
    filename: str, content: bytes
) -> tuple[list[dict[str, object]], dict[str, int]]:
    """Parse every row and return both the accepted rows and a drop-reason
    histogram. The extraction agent reports the histogram back to the CRM
    so users see exactly why a file ingested fewer contacts than expected.
    """
    rows: list[dict[str, object]] = []
    stats: dict[str, int] = {
        "total_rows_read": 0,
        "rows_accepted": 0,
        "rows_missing_email": 0,
        "rows_invalid_email": 0,
        "rows_custom_only": 0,
        "rows_missing_name": 0,
        "rows_missing_company": 0,
    }
    name = filename.lower()

    def _process(raw_row: dict[str, str]) -> None:
        stats["total_rows_read"] += 1
        normalized, reason = _normalize_row_with_reason(raw_row)
        if normalized is not None:
            # Strip the internal sentinel before handing to the upsert
            # pipeline — it's only for stats accounting.
            if normalized.pop("_missing_name", False):
                stats["rows_missing_name"] += 1
            if not normalized.get("company"):
                stats["rows_missing_company"] += 1
            rows.append(normalized)
            stats["rows_accepted"] += 1
        elif reason == "no_email":
            stats["rows_missing_email"] += 1
        elif reason == "invalid_email":
            stats["rows_invalid_email"] += 1
        elif reason == "custom_only":
            stats["rows_custom_only"] += 1

    if name.endswith(".csv"):
        reader = csv.DictReader(
            io.TextIOWrapper(io.BytesIO(content), encoding="utf-8", newline="")
        )
        for raw in reader:
            _process(raw)
    elif name.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is not None:
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return rows, stats
            headers = [str(h) if h is not None else "" for h in header_row]
            for row in rows_iter:
                raw = {
                    headers[i]: ("" if row[i] is None else str(row[i]))
                    for i in range(len(headers))
                    if i < len(row)
                }
                _process(raw)
    elif name.endswith(".json"):
        import ijson

        bio = io.BytesIO(content)
        head = bio.read(1)
        remainder = bio.read()
        combined = io.BytesIO(head + remainder)
        if head == b"[":
            prefix = "item"
        elif head == b"{":
            prefix = "contacts.item"
        else:
            return rows, stats
        for raw in ijson.items(combined, prefix):
            if not isinstance(raw, dict):
                continue
            _process({k: ("" if v is None else str(v)) for k, v in raw.items()})
    else:
        raise ValueError(f"Unsupported file type: {filename}")

    return rows, stats
